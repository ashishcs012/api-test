import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_student():
    st_add_url ="http://thetestingworldapi.com/api/studentsDetails"
    #f = open('C:\\python-t\\API-testing\\utilities\\studentdata.xlsx')
    wb = openpyxl.load_workbook("C:\\python-t\\API-testing\\utilities\\studentdata.xlsx")
    sh = wb['Sheet1']
    rows = sh.max_row
    print(sh)

    for i in range(1,rows+1):
        cell_first_name = sh.cell(row=i,column=1)
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_date_of_birth = sh.cell(row=i, column=4)



